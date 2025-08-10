#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import requests
import datetime
import logging
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    filters, ConversationHandler, ContextTypes
)
from openpyxl import Workbook, load_workbook

# ===== Конфиг =====
BOT_TOKEN = os.environ.get("BOT_TOKEN", "")  # обязателен в Render/Env
ADMIN_ID = int(os.environ.get("ADMIN_ID", "0"))  # ваш Telegram ID (число)
FILE_NAME = "calculations.xlsx"

# Состояния для ConversationHandler
PRICE, ENGINE, YEAR, DELIVERY, FREIGHT, BROKER = range(6)

# Настройка логов
logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

if not BOT_TOKEN:
    logger.error("BOT_TOKEN не задан в переменных окружения. Бот не запустится.")
    # не выходим, чтобы при тестировании сообщить об ошибке в логах

# ===== Утилиты =====
def get_jpy_rate():
    """
    Берёт курс JPY (1 йена в рублях) с ЦБ РФ.
    Возвращает float или None в случае ошибки.
    """
    urls = [
        "https://www.cbr-xml-daily.ru/daily_json.js",
        "https://www.cbr.ru/scripts/XML_daily.asp"  # резерв (не json) - не используем по умолчанию
    ]
    try:
        r = requests.get(urls[0], timeout=8)
        r.raise_for_status()
        data = r.json()
        val = data.get("Valute", {}).get("JPY", {}).get("Value")
        if val is None:
            logger.warning("Не найден JPY в ответе ЦБ.")
            return None
        return float(val)
    except Exception as e:
        logger.exception("Ошибка получения курса ЦБ: %s", e)
        return None

def save_to_excel(row):
    """
    row - список значений в том порядке, который используем в шапке.
    Если файла нет — создаёт и добавляет шапку.
    """
    header = [
        "Дата", "Цена ¥", "Курс ¥", "Цена ₽",
        "Двигатель (см³)", "Год выпуска", "Возраст (лет)",
        "Доставка ₽", "Фрахт ₽", "Пошлина ₽", "Утильсбор ₽",
        "НДС ₽", "Брокер ₽", "Итого ₽"
    ]
    try:
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            wb.save(FILE_NAME)

        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append(row)
        wb.save(FILE_NAME)
    except Exception as e:
        logger.exception("Ошибка при сохранении в Excel: %s", e)

def calc_customs(price_rub, age, engine_cc):
    """
    Пример расчёта пошлины, как обсуждали выше.
    Возвращает tuple: (duty, util_fee, vat)
    Внимание: ставки примерные — при необходимости можно изменить.
    """
    if age < 3:
        duty = price_rub * 0.48  # 48% от таможенной стоимости — пример
    else:
        cc = engine_cc
        if cc <= 1000:
            duty = cc * 1.5 * 100
        elif cc <= 1500:
            duty = cc * 1.7 * 100
        elif cc <= 1800:
            duty = cc * 2.5 * 100
        elif cc <= 2300:
            duty = cc * 2.7 * 100
        elif cc <= 3000:
            duty = cc * 3.0 * 100
        else:
            duty = cc * 3.6 * 100

    util_fee = 5200 if age < 3 else 8200
    vat = (price_rub + duty) * 0.2  # 20% НДС
    return duty, util_fee, vat

# ===== HANDLERS =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Начало диалога. Получаем курс и просим ввести цену в йенах.
    """
    rate = get_jpy_rate()
    if rate:
        context.user_data['rate'] = rate
        await update.message.reply_text(
            f"📈 Актуальный курс: 1 ¥ = {rate:.4f} ₽\n"
            "Введите цену на аукционе (в йенах), например: 1200000"
        )
        return PRICE
    else:
        # если не удалось получить курс, всё равно даём вводить цену, но сообщаем
        await update.message.reply_text(
            "⚠️ Не удалось получить курс йены с ЦБ. Введите цену на аукционе (в йенах),\n"
            "и затем в следующем шаге введите курс вручную (в рублях за 1 ¥)."
        )
        context.user_data['rate'] = None
        return PRICE

async def price_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", "").replace(" ", "")
    try:
        price_yen = float(text)
        context.user_data['price_yen'] = price_yen
        # если курса нет — попросим ввести
        if context.user_data.get('rate') is None:
            await update.message.reply_text("Введи курс йены к рублю (например, 0.6123):")
            return YEAR  # временно используем слот YEAR для ввода курса
        else:
            await update.message.reply_text("Введи объём двигателя (в см³), например: 1798")
            return ENGINE
    except ValueError:
        await update.message.reply_text("Пожалуйста, введи число — цену в йенах.")
        return PRICE

async def year_as_rate_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Этот хендлер используется только если мы не получили курс автоматически.
    Пользователь ввёл курс вручную. Далее попросим мотор/год.
    """
    text = update.message.text.strip().replace(",", ".")
    try:
        rate = float(text)
        context.user_data['rate'] = rate
        await update.message.reply_text("Введи объём двигателя (в см³), например: 1798")
        return ENGINE
    except ValueError:
        await update.message.reply_text("Пожалуйста, введи число — курс в формате 0.6123")
        return YEAR

async def engine_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        cc = int(text)
        context.user_data['engine_cc'] = cc
        await update.message.reply_text("Введи год выпуска автомобиля (например, 2018):")
        return YEAR
    except ValueError:
        await update.message.reply_text("Пожалуйста, введи целое число в см³ (например, 1798).")
        return ENGINE

async def year_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        year = int(text)
        current_year = datetime.datetime.now().year
        if year > current_year or year < 1980:
            await update.message.reply_text(f"Некорректный год. Введи год от 1980 до {current_year}.")
            return YEAR
        context.user_data['year'] = year
        context.user_data['age'] = current_year - year
        await update.message.reply_text("Введи стоимость доставки в порт Японии (в рублях), например: 25000")
        return DELIVERY
    except ValueError:
        await update.message.reply_text("Пожалуйста, введи год выпуска числом (например, 2018).")
        return YEAR

async def delivery_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".").replace(" ", "")
    try:
        context.user_data['delivery'] = float(text)
        await update.message.reply_text("Введи стоимость фрахта до России (в рублях), например: 45000")
        return FREIGHT
    except ValueError:
        await update.message.reply_text("Пожалуйста, введи число (рубли).")
        return DELIVERY

async def freight_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".").replace(" ", "")
    try:
        context.user_data['freight'] = float(text)
        await update.message.reply_text("Введи услуги брокера (в рублях), например: 15000")
        return BROKER
    except ValueError:
        await update.message.reply_text("Пожалуйста, введи число (рубли).")
        return FREIGHT

async def broker_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".").replace(" ", "")
    try:
        broker_fee = float(text)
        ctx = context.user_data
        # основная математика
        price_yen = ctx['price_yen']
        rate = ctx['rate']
        price_rub = price_yen * rate  # цена в рублях

        duty, util_fee, vat = calc_customs(price_rub, ctx['age'], ctx['engine_cc'])

        delivery = ctx.get('delivery', 0.0)
        freight = ctx.get('freight', 0.0)
        total = price_rub + delivery + freight + duty + util_fee + vat + broker_fee

        # Сохранение в Excel
        save_row = [
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            price_yen, rate, round(price_rub, 2),
            ctx['engine_cc'], ctx['year'], ctx['age'],
            delivery, freight, round(duty, 2), util_fee, round(vat, 2),
            broker_fee, round(total, 2)
        ]
        save_to_excel(save_row)

        # Формируем подробный ответ
        breakdown = (
            "📊 *Подробный расчёт:*\n\n"
            f"💴 Цена на аукционе: {price_yen:,.0f} ¥ × {rate:.4f} ₽ = {price_rub:,.0f} ₽\n"
            f"🚢 Доставка в порт Японии: {delivery:,.0f} ₽\n"
            f"⛴ Фрахт до России: {freight:,.0f} ₽\n"
            f"💼 Пошлина: {duty:,.0f} ₽\n"
            f"♻️ Утильсбор: {util_fee:,.0f} ₽\n"
            f"💰 НДС (20%): {vat:,.0f} ₽\n"
            f"📑 Услуги брокера: {broker_fee:,.0f} ₽\n\n"
            f"📅 Возраст авто: {ctx['age']} лет\n"
            f"💵 *Итого*: {total:,.0f} ₽"
        )

        await update.message.reply_text(breakdown, parse_mode="Markdown")
        return ConversationHandler.END

    except ValueError:
        await update.message.reply_text("Пожалуйста, введи число (рубли).")
        return BROKER
    except Exception as e:
        logger.exception("Ошибка в broker_input: %s", e)
        await update.message.reply_text("Произошла ошибка при расчёте. Попробуй ещё раз.")
        return ConversationHandler.END

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Отправляет calculations.xlsx только ADMIN_ID.
    """
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("⛔ У вас нет доступа к этой команде.")
        return

    if not os.path.exists(FILE_NAME):
        await update.message.reply_text("Файл с расчётами ещё не создан.")
        return

    try:
        # файл отправляем как документ
        await context.bot.send_document(chat_id=update.effective_chat.id, document=open(FILE_NAME, "rb"))
    except Exception as e:
        logger.exception("Ошибка отправки Excel: %s", e)
        await update.message.reply_text("Ошибка при отправке файла. Попробуйте позже.")

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Расчёт отменён.")
    return ConversationHandler.END

# ===== Main =====
def main():
    if not BOT_TOKEN:
        logger.error("BOT_TOKEN не задан — завершение.")
        return

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, price_input)],
            ENGINE: [MessageHandler(filters.TEXT & ~filters.COMMAND, engine_input)],
            YEAR: [
                MessageHandler(filters.Regex(r'^\d+(\.\d+)?$') & ~filters.COMMAND, year_as_rate_input),  # если ввод курса (когда rate не был загружен)
                MessageHandler(filters.TEXT & ~filters.COMMAND, year_input)
            ],
            DELIVERY: [MessageHandler(filters.TEXT & ~filters.COMMAND, delivery_input)],
            FREIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, freight_input)],
            BROKER: [MessageHandler(filters.TEXT & ~filters.COMMAND, broker_input)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True
    )

    app.add_handler(conv_handler)
    app.add_handler(CommandHandler("export", export_excel))

    logger.info("Запуск бота...")
    app.run_polling()

if __name__ == "__main__":
    main()
